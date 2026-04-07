"""
Excel Import/Export Handler for Akhgam Herbals
Uses openpyxl for reading/writing Excel files
"""
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from werkzeug.security import generate_password_hash


def _style_header(ws, headers):
    """Apply styled headers to worksheet."""
    header_font = Font(name='Poppins', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2D6A4F', end_color='2D6A4F', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def export_products(products, output_dir):
    """
    Export products list (list of dicts) to a styled Excel file.
    Returns the file path.
    """
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, 'akhgam_products.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'

    headers = ['ID', 'Name', 'Category', 'Benefits', 'Description',
               'Price (₹)', 'Original Price (₹)', 'Image', 'Status',
               'Featured', 'Rating', 'Reviews Count', 'Created At']

    _style_header(ws, headers)

    # Column widths
    widths = [6, 35, 15, 40, 50, 12, 15, 20, 10, 10, 8, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    data_align = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row_num, p in enumerate(products, 2):
        values = [
            p['id'],
            p['name'],
            p.get('category', ''),
            p.get('benefits', ''),
            p.get('description', ''),
            float(p.get('price', 0)),
            float(p.get('original_price', 0)) if p.get('original_price') else '',
            p.get('image', 'default.jpg'),
            p.get('status', 'active'),
            'Yes' if p.get('featured') else 'No',
            float(p.get('rating', 0)),
            int(p.get('reviews_count', 0)),
            str(p.get('created_at', ''))
        ]
        for col_num, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.alignment = data_align
            cell.border = thin_border

        # Alternate row coloring
        if row_num % 2 == 0:
            fill = PatternFill(start_color='F0FAF4', end_color='F0FAF4', fill_type='solid')
            for col_num in range(1, len(values) + 1):
                ws.cell(row=row_num, column=col_num).fill = fill

    wb.save(filepath)
    return filepath


def export_users(users, output_dir):
    """
    Export users list (list of dicts) to a styled Excel file.
    Returns the file path.
    """
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, 'akhgam_users.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Clients'

    headers = ['ID', 'Name', 'Email', 'Phone', 'Age', 'Gender', 'Address', 'State', 'District', 'Pincode', 'Role', 'Status', 'Joined']

    _style_header(ws, headers)

    widths = [6, 25, 30, 15, 6, 10, 35, 16, 16, 10, 10, 10, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    data_align = Alignment(vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row_num, u in enumerate(users, 2):
        values = [
            u['id'],
            u['name'],
            u['email'],
            u.get('phone', ''),
            u.get('age', '') or '',
            (u.get('gender', '') or '').capitalize(),
            u.get('address', '') or '',
            u.get('state', '') or '',
            u.get('district', '') or '',
            u.get('pincode', '') or '',
            u.get('role', 'client'),
            u.get('status', 'active'),
            str(u.get('created_at', ''))
        ]
        for col_num, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.alignment = data_align
            cell.border = thin_border

        if row_num % 2 == 0:
            fill = PatternFill(start_color='F0FAF4', end_color='F0FAF4', fill_type='solid')
            for col_num in range(1, len(values) + 1):
                ws.cell(row=row_num, column=col_num).fill = fill

    wb.save(filepath)
    return filepath


def import_products(filepath, cursor, db):
    """
    Import products from an Excel file.
    Expected columns: Name, Category, Benefits, Description, Price,
                      Original Price, Status, Featured, Rating, Reviews Count
    Returns count of imported rows.
    """
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    count = 0

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or not row[0]:
            continue

        # Skip the ID column if present (auto-detect by checking if first col is numeric)
        offset = 0
        if isinstance(row[0], (int, float)):
            offset = 1

        name = str(row[0 + offset] or '').strip()
        if not name:
            continue

        category = str(row[1 + offset] or 'General').strip()
        benefits = str(row[2 + offset] or '').strip()
        description = str(row[3 + offset] or '').strip()
        price = float(row[4 + offset] or 0)
        original_price = float(row[5 + offset] or 0) if row[5 + offset] else None

        # Handle optional columns
        try:
            image = str(row[6 + offset] or 'default.jpg').strip()
        except (IndexError, TypeError):
            image = 'default.jpg'

        try:
            status = str(row[7 + offset] or 'active').strip().lower()
            if status not in ('active', 'inactive'):
                status = 'active'
        except (IndexError, TypeError):
            status = 'active'

        try:
            featured_val = row[8 + offset]
            featured = 1 if featured_val and str(featured_val).strip().lower() in ('1', 'yes', 'true') else 0
        except (IndexError, TypeError):
            featured = 0

        try:
            rating = float(row[9 + offset] or 0)
        except (IndexError, TypeError, ValueError):
            rating = 0

        try:
            reviews_count = int(row[10 + offset] or 0)
        except (IndexError, TypeError, ValueError):
            reviews_count = 0

        if price <= 0:
            continue

        cursor.execute(
            """INSERT INTO products (name, category, benefits, description, price,
               original_price, image, status, featured, rating, reviews_count)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (name, category, benefits, description, price,
             original_price, image, status, featured, rating, reviews_count)
        )
        count += 1

    db.commit()
    wb.close()
    return count


def import_users(filepath, cursor, db):
    """
    Import users from an Excel file.
    Expected columns: Name, Email, Phone, Age, Gender, Address, State, District, Pincode
    Password defaults to 'password123'.
    Returns count of imported rows.
    """
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    count = 0
    default_password = generate_password_hash('password123')

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or not row[0]:
            continue

        # Skip the ID column if present
        offset = 0
        if isinstance(row[0], (int, float)):
            offset = 1

        name = str(row[0 + offset] or '').strip()
        if not name:
            continue

        email = str(row[1 + offset] or '').strip()
        if not email:
            continue

        phone = str(row[2 + offset] or '').strip()

        # Optional fields: age, gender, address, state, district, pincode
        try:
            age = int(row[3 + offset]) if row[3 + offset] else None
        except (IndexError, TypeError, ValueError):
            age = None

        try:
            gender = str(row[4 + offset] or '').strip().lower()
            if gender not in ('male', 'female', 'other'):
                gender = None
        except (IndexError, TypeError):
            gender = None

        try:
            address = str(row[5 + offset] or '').strip() or None
        except (IndexError, TypeError):
            address = None

        try:
            pincode = str(row[6 + offset] or '').strip() or None
        except (IndexError, TypeError):
            pincode = None

        # Support both old format (without state/district) and new format
        state = None
        district = None
        if len(row) >= (12 + offset):
            try:
                state = str(row[6 + offset] or '').strip() or None
            except (IndexError, TypeError):
                state = None
            try:
                district = str(row[7 + offset] or '').strip() or None
            except (IndexError, TypeError):
                district = None
            try:
                pincode = str(row[8 + offset] or '').strip() or None
            except (IndexError, TypeError):
                pincode = None

        # Check if user already exists
        cursor.execute("SELECT id FROM users WHERE email=%s", (email,))
        if cursor.fetchone():
            continue

        cursor.execute(
            "INSERT INTO users (name, email, phone, age, gender, address, state, district, pincode, password, role) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'client')",
            (name, email, phone, age, gender, address, state, district, pincode, default_password)
        )
        count += 1

    db.commit()
    wb.close()
    return count
